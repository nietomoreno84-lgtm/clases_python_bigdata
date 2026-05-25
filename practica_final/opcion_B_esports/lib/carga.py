import csv
from openpyxl import load_workbook
import json


def cargar_csv(carpeta, nombre):
    """
    Carga un archivo CSV y devuelve su contenido como una lista de diccionarios.

    Cada fila del archivo CSV se convierte en un diccionario donde:
    - Las claves corresponden a los nombres de las columnas.
    - Los valores corresponden a los datos de cada fila.

    Parámetros:
        carpeta (str): Ruta de la carpeta donde se encuentra el archivo CSV.
        fichero (str): Nombre del archivo CSV a cargar.

    Retorna:
        list[dict]: Lista de diccionarios con los datos del CSV.

    Ejemplo:
        >>> datos = cargar_csv("datos", "usuarios.csv")
        >>> print(datos[0])
        {'nombre': 'Ana', 'edad': '25', 'ciudad': 'Madrid'}

    Requisitos:
        - El módulo `csv` debe estar importado previamente.
        - El archivo debe existir y estar codificado en UTF-8.

    Excepciones:
        FileNotFoundError:
            Se produce si el archivo no existe.
        UnicodeDecodeError:
            Se produce si el archivo no está codificado en UTF-8.
    """
    fichero = open(f"{carpeta}/{nombre}", "r", encoding='UTF-8')
    lector = csv.DictReader(fichero)
    lista = list(lector)
    pintar_datos_ficheros(nombre,lista)
    fichero.close()
    return lista



def cargar_excel(carpeta, nombre, hoja2=None):
    """
    Carga un archivo Excel y devuelve su contenido como una lista de diccionarios.

    La función utiliza la primera fila de la hoja activa como cabecera y
    transforma cada fila restante en un diccionario donde:
    - Las claves son los nombres de las columnas.
    - Los valores son los datos correspondientes de cada fila.

    Parámetros:
        carpeta (str): Ruta de la carpeta donde se encuentra el archivo Excel.
        fichero (str): Nombre del archivo Excel (.xlsx) que se desea cargar.

    Retorna:
        list[dict]: Lista de diccionarios con la información contenida
                    en el archivo Excel.

    Funcionamiento:
        - Abre el archivo Excel indicado.
        - Accede a la hoja activa del documento.
        - Obtiene las cabeceras desde la primera fila.
        - Recorre el resto de filas y crea un diccionario por cada una.
        - Devuelve una lista con todos los registros.

    Ejemplo:
        Supongamos un archivo `jugadores.xlsx` con el contenido:

            nombre      posicion    edad
            Carlos      Delantero   24
            Ana         Portera     27

        >>> datos = cargar_excel("datos", "jugadores.xlsx")
        >>> print(datos)

        [
            {
                'nombre': 'Carlos',
                'posicion': 'Delantero',
                'edad': 24
            },
            {
                'nombre': 'Ana',
                'posicion': 'Portera',
                'edad': 27
            }
        ]

    Requisitos:
        - Tener instalada la librería `openpyxl`.
        - Haber importado previamente:
              from openpyxl import load_workbook
        - El archivo debe existir y ser un Excel válido.

    Excepciones:
        FileNotFoundError:
            Se produce si el archivo no existe.
        PermissionError:
            Se produce si el archivo está siendo utilizado por otro programa.
        InvalidFileException:
            Se produce si el archivo no tiene un formato Excel válido.
    """
    excel = load_workbook(f'./{carpeta}/{nombre}')
    if hoja2:
        hoja=excel[hoja2]
    else:    
         hoja = excel.active
    lista = []

   
    filas = hoja.iter_rows(values_only = True)
    cabeceras = next(filas)

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        diccionario = dict(zip(cabeceras, fila))
        lista.append(diccionario)

    pintar_datos_ficheros(nombre,lista)
 
    return lista



def cargar_json(carpeta, nombre):
    """
    Carga un archivo JSON y devuelve su contenido.

    La función abre un archivo JSON ubicado en la carpeta indicada,
    lee su contenido y lo convierte en estructuras de datos de Python
    utilizando `json.load()`.

    Parámetros:
        carpeta (str): Ruta de la carpeta donde se encuentra el archivo JSON.
        nombre (str): Nombre del archivo JSON a cargar.

    Retorna:
        dict | list:
            Contenido del archivo JSON convertido a estructuras de Python.
            Dependiendo del JSON, puede devolver:
            - Un diccionario (`dict`)
            - Una lista (`list`)

    Funcionamiento:
        - Abre el archivo en modo lectura con codificación UTF-8.
        - Convierte el contenido JSON en objetos Python.
        - Devuelve los datos cargados.
        - Si el archivo no existe, muestra un mensaje de error.

    Ejemplo:
        Supongamos un archivo `usuarios.json` con el contenido:

        {
            "nombre": "Ana",
            "edad": 25
        }

        >>> datos = cargar_json("datos", "usuarios.json")
        >>> print(datos)

        {
            'nombre': 'Ana',
            'edad': 25
        }

    Requisitos:
        - Haber importado previamente el módulo `json`:
              import json
        - El archivo debe tener formato JSON válido.
        - El archivo debe estar codificado en UTF-8.

    Excepciones:
        FileNotFoundError:
            Se produce si el archivo o la carpeta no existen.
        JSONDecodeError:
            Se produce si el contenido del archivo no es un JSON válido.

    Notas:
        - En caso de que el archivo no exista, la función muestra
          el mensaje:
              'Archivo o carpeta no encontrado'
        - En ese caso, la función devuelve `None`.
    """   
    try:
        fichero = open(f"./{carpeta}/{nombre}", "r", encoding="UTF-8")
        datos = json.load(fichero)

        pintar_datos_ficheros(nombre,datos)

        return datos
    except FileNotFoundError:
        print('Archivo o carpeta no encontrado')


def pintar_datos_ficheros(fichero,lista):
    print('Numero total de registros:', len(lista))
    print('Nombre del fichero', fichero)
    print('Nombre de los campos:', list(lista[0].keys()))
    for i in range(5):
            print(lista[i])
            

    