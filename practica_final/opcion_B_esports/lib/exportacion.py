import csv
import os
from openpyxl import Workbook, load_workbook



def crear_csv(lista, nombre, carpeta):
    """
    Crea un archivo CSV y escribe en él los datos de una lista de diccionarios.

    La función genera un archivo CSV utilizando las claves del primer
    diccionario como cabeceras y escribe todos los registros contenidos
    en la lista.

    Parámetros:
        lista (list[dict]):
            Lista de diccionarios con los datos que se desean guardar
            en el archivo CSV.

        nombre (str):
            Nombre del archivo CSV que se va a crear.

        carpeta (str):
            Ruta de la carpeta donde se guardará el archivo.

    Retorna:
        None

    Funcionamiento:
        - Abre o crea el archivo CSV en modo escritura/añadido.
        - Obtiene automáticamente las cabeceras desde el primer elemento.
        - Crea un objeto `csv.DictWriter`.
        - Escribe las cabeceras en la primera fila.
        - Escribe todos los registros de la lista.
        - Cierra el archivo correctamente.

    Ejemplo:
        >>> jugadores = [
        ...     {
        ...         "nombre": "Carlos",
        ...         "equipo": "Dragons",
        ...         "edad": 22
        ...     },
        ...     {
        ...         "nombre": "Ana",
        ...         "equipo": "Titans",
        ...         "edad": 25
        ...     }
        ... ]

        >>> crear_csv(jugadores, "jugadores.csv", "datos")

        Resultado:
            Se crea el archivo:
            ./datos/jugadores.csv

            Con el contenido:

            nombre,equipo,edad
            Carlos,Dragons,22
            Ana,Titans,25

    Requisitos:
        - Haber importado previamente el módulo `csv`:
              import csv
        - La carpeta destino debe existir.
        - La lista debe contener al menos un diccionario.

    Excepciones:
        FileNotFoundError:
            Se produce si la carpeta no existe.

        PermissionError:
            Se produce si no hay permisos de escritura.

        IndexError:
            Se produce si la lista está vacía y se intenta acceder
            al primer elemento.

    Notas:
        - Las cabeceras se generan automáticamente a partir de las claves
          del primer diccionario de la lista.
        - El archivo se abre en modo `'a'`, por lo que añadirá contenido
          al archivo si ya existe.
        - `newline=''` evita líneas vacías adicionales en Windows.
    """
    fichero = open(f"./{carpeta}/{nombre}", 'a', encoding='UTF-8', newline='')
    # saber las cabeceras.
    cabeceras = []
    for key in lista[0].keys():
        cabeceras.append(key)
    #volcar los datos de la lista en un objeto csv
    mi_csv = csv.DictWriter(fichero, fieldnames=cabeceras) 
    #imprimir cabeceras en la primera linea
    mi_csv.writeheader()
    # Escribir cada fila en el csv
    mi_csv.writerows(lista)
    # cerramos el fichero
    fichero.close()


def crear_excel(carpeta, fichero, datos, hoja):
    """
    Crea un archivo Excel o añade una nueva hoja con datos.

    La función permite generar un archivo Excel utilizando una lista
    de diccionarios. Si el archivo ya existe, se carga y se añade una
    nueva hoja; si no existe, se crea un nuevo libro de Excel.

    Parámetros:
        carpeta (str):
            Ruta de la carpeta donde se guardará el archivo Excel.

        fichero (str):
            Nombre del archivo Excel (.xlsx).

        datos (list[dict]):
            Lista de diccionarios con la información que se escribirá
            en la hoja de cálculo.

        hoja (str):
            Nombre de la hoja que se creará dentro del archivo Excel.

    Retorna:
        None

    Funcionamiento:
        - Comprueba si el archivo Excel ya existe.
        - Si existe, carga el libro actual.
        - Si no existe, crea un nuevo libro.
        - Crea una nueva hoja con el nombre indicado.
        - Obtiene automáticamente las cabeceras desde el primer diccionario.
        - Escribe las cabeceras en la primera fila.
        - Recorre los datos y añade cada registro como una fila.
        - Guarda el archivo Excel en la ruta indicada.

    Ejemplo:
        >>> jugadores = [
        ...     {
        ...         "nombre": "Carlos",
        ...         "equipo": "Dragons",
        ...         "edad": 22
        ...     },
        ...     {
        ...         "nombre": "Ana",
        ...         "equipo": "Titans",
        ...         "edad": 25
        ...     }
        ... ]

        >>> crear_excel(
        ...     carpeta="datos",
        ...     fichero="jugadores.xlsx",
        ...     datos=jugadores,
        ...     hoja="Temporada 2026"
        ... )

        Resultado:
            Se crea el archivo:
            ./datos/jugadores.xlsx

            Con una hoja llamada:
            "Temporada 2026"

    Requisitos:
        - Tener instalada la librería `openpyxl`.
        - Haber importado previamente:
              from openpyxl import Workbook, load_workbook
        - Haber importado:
              import os
        - La carpeta destino debe existir.
        - La lista `datos` debe contener al menos un elemento.

    Excepciones:
        FileNotFoundError:
            Se produce si la carpeta no existe.

        PermissionError:
            Se produce si el archivo está abierto o bloqueado.

        IndexError:
            Se produce si la lista `datos` está vacía.

        KeyError:
            Puede producirse si los diccionarios no tienen las mismas claves.

    Notas:
        - Las cabeceras se generan automáticamente utilizando las claves
          del primer diccionario.
        - Los valores se insertan respetando el orden de las cabeceras.
        - Cada llamada crea una nueva hoja dentro del libro Excel.
    """
    ruta = f"./{carpeta}/{fichero}"
    if os.path.exists(ruta):
    # Cargarlo si existe
        wb = load_workbook(ruta)
    else:
        # Crear uno nuevo si no existe
        wb = Workbook()
    
    ws =  wb.create_sheet(title=hoja)
        
       
        # crear primero el libro de excel
        
    # extraer de un diccionario cualquier de mi lista las cabeceras.
    cabeceras = list(datos[0].keys())
    # quiero añadirlo a mi hoja
    ws.append(cabeceras)
    
    # recorremos nuestra de lista de datos para imprimir en cada fila un dato concreto
    for item in datos:
        # para que esto funcione el empleado tiene que tener los datos en el mismo orden que la lista caberas. Y estar convertido en lista.
        # lista_empleado = list(empleado.values())
        lista_item = [item[clave] for clave in cabeceras ]
        ws.append(lista_item)
    
    wb.save(f'./{carpeta}/{fichero}')


def generar_informe_txt(diccionario_auditoria, diccionario_cambios, ruta_txt):
    """
    Genera el archivo de texto estructurado con el resumen del proceso.
    """
    with open(ruta_txt, mode="w", encoding="utf-8") as archivo:
        archivo.write("==================================================\n")
        archivo.write("         INFORME DE LIMPIEZA - CYBERLEAGUE        \n")
        archivo.write("==================================================\n\n")
        
        # Recorrer los resultados para escribir las métricas por cada archivo
        for nombre_fichero, datos in diccionario_auditoria.items():
            archivo.write(f"FICHERO: {nombre_fichero}\n")
            archivo.write(f"--------------------------------------------------\n")
            archivo.write(f"Total de registros procesados: {datos['total_registros']}\n")
            archivo.write(f"Registros duplicados eliminados: {datos['duplicados']}\n")
            
            # Sumar el total de celdas vacías encontradas en este archivo
            total_vacios = sum(datos['valores_vacios'].values())
            archivo.write(f"  • Total de valores vacíos gestionados: {total_vacios}\n")
            
            # Mostrar los cambios cuantificados reales que aportó tu otra función
            cambios_reales = diccionario_cambios.get(nombre_fichero, 0)
            archivo.write(f"  • Total de celdas modificadas/corregidas: {cambios_reales}\n\n")
            
        archivo.write("==================================================\n")
        archivo.write("Fin del informe.\n")
        
    print(f"Fichero de texto creado: {ruta_txt}")
